import openpyxl


def readData(file, sheet_name, ro, col):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.Cell(row=ro, column=col).value


def writeData(file, sheet_name, ro, col, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row=ro, column=col).value = data
    workbook.save(file)


def rowCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row()
